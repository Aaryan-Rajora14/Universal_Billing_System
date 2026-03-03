import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

class UniversalExcelManager:
    def __init__(self, filename='store_bills.xlsx'):
        self.filename = filename
        self.init_excel_file()

    def init_excel_file(self):
        if not os.path.exists(self.filename):
            wb = Workbook()

            # ----- Bills Sheet -----
            ws_bills = wb.active
            ws_bills.title = "Bills"
            headers_bills = [
                "Sr No", "Date", "Customer Name", "Customer ID", "Bill No",
                "Items Count", "Subtotal (₹)", "Discount (₹)", "Grand Total (₹)", "QR Code", "PDF File"
            ]
            for col, header in enumerate(headers_bills, 1):
                cell = ws_bills.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            col_widths = {'A': 8, 'B': 12, 'C': 22, 'D': 15, 'E': 22,
                          'F': 10, 'G': 15, 'H': 15, 'I': 15, 'J': 20, 'K': 20}
            for col, width in col_widths.items():
                ws_bills.column_dimensions[col].width = width

            # ----- Profit & Loss Sheet -----
            ws_pl = wb.create_sheet("ProfitLoss")
            headers_pl = ["Bill No", "Date", "Subtotal (₹)", "Discount (₹)", "Grand Total (₹)", "Profit (₹)"]
            for col, header in enumerate(headers_pl, 1):
                cell = ws_pl.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            col_widths_pl = {'A': 22, 'B': 12, 'C': 15, 'D': 15, 'E': 15, 'F': 15}
            for col, width in col_widths_pl.items():
                ws_pl.column_dimensions[col].width = width

            # ----- Settings Sheet -----
            ws_settings = wb.create_sheet("Settings")
            ws_settings.append(["Setting", "Value"])
            ws_settings.append(["store_name", "My Store"])
            ws_settings.append(["store_tagline", "Quality You Can Trust"])
            ws_settings.append(["store_address", "123 Main Street, City - 000000"])
            ws_settings.append(["store_phone", "+91 98765 43210"])
            ws_settings.append(["store_email", "contact@mystore.com"])
            ws_settings.append(["store_website", "www.mystore.com"])
            ws_settings.append(["currency_symbol", "₹"])
            ws_settings.append(["profit_margin", 30])
            ws_settings.append(["customer_counter", 1000])
            ws_settings.append(["bill_counter", 1])
            for col in range(1, 3):
                cell = ws_settings.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            ws_settings.column_dimensions['A'].width = 25
            ws_settings.column_dimensions['B'].width = 40

            wb.save(self.filename)

    # ---------- Settings ----------
    def _get_setting(self, key, default=''):
        wb = load_workbook(self.filename)
        ws = wb["Settings"]
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 1).value == key:
                val = ws.cell(row, 2).value
                return val if val is not None else default
        return default

    def _set_setting(self, key, value):
        wb = load_workbook(self.filename)
        ws = wb["Settings"]
        found = False
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 1).value == key:
                ws.cell(row, 2).value = value
                found = True
                break
        if not found:
            ws.append([key, value])
        wb.save(self.filename)

    def get_store_settings(self):
        return {
            'store_name': self._get_setting('store_name', 'My Store'),
            'store_tagline': self._get_setting('store_tagline', 'Quality You Can Trust'),
            'store_address': self._get_setting('store_address', ''),
            'store_phone': self._get_setting('store_phone', ''),
            'store_email': self._get_setting('store_email', ''),
            'store_website': self._get_setting('store_website', ''),
            'currency_symbol': self._get_setting('currency_symbol', '₹'),
            'profit_margin': float(self._get_setting('profit_margin', 30)),
        }

    def update_store_settings(self, settings_dict):
        for key, value in settings_dict.items():
            self._set_setting(key, value)

    def get_profit_margin(self):
        return float(self._get_setting('profit_margin', 30))

    def set_profit_margin(self, value):
        self._set_setting('profit_margin', value)

    def get_currency_symbol(self):
        return self._get_setting('currency_symbol', '₹')

    # ---------- Customer ID Management ----------
    def get_or_create_customer_id(self, customer_name):
        if not customer_name:
            return "CUST-0000"
        wb = load_workbook(self.filename)
        ws_bills = wb["Bills"]
        customer_ids = {}
        for row in range(2, ws_bills.max_row + 1):
            name = ws_bills.cell(row, 3).value
            cid = ws_bills.cell(row, 4).value
            if name and cid:
                customer_ids[name.strip().lower()] = cid
        if customer_name.strip().lower() in customer_ids:
            return customer_ids[customer_name.strip().lower()]
        ws_settings = wb["Settings"]
        counter = 1000
        for row in range(2, ws_settings.max_row + 1):
            if ws_settings.cell(row, 1).value == "customer_counter":
                counter = int(ws_settings.cell(row, 2).value) + 1
                ws_settings.cell(row, 2).value = counter
                break
        wb.save(self.filename)
        return f"CUST-{counter}"

    # ---------- Bill Number ----------
    def generate_bill_no(self):
        wb = load_workbook(self.filename)
        ws = wb["Settings"]
        counter = 1
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 1).value == "bill_counter":
                counter = int(ws.cell(row, 2).value)
                ws.cell(row, 2).value = counter + 1
                break
        wb.save(self.filename)
        date = datetime.now().strftime("%Y%m%d")
        return f"BILL-{date}-{counter:04d}"

    # ---------- Save Bill ----------
    def save_bill_data(self, customer_name, customer_id, bill_no, items,
                       subtotal, discount_total, grand_total, qr_file, pdf_file):
        wb = load_workbook(self.filename)
        ws_bills = wb["Bills"]
        ws_pl = wb["ProfitLoss"]

        sr_no = ws_bills.max_row
        now = datetime.now()
        date_str = now.strftime("%d-%m-%Y")
        items_count = len(items)
        row_data = [
            sr_no, date_str, customer_name, customer_id, bill_no,
            items_count, subtotal, discount_total, grand_total, qr_file, pdf_file
        ]
        ws_bills.append(row_data)

        last_row = ws_bills.max_row
        for col in range(1, 12):
            cell = ws_bills.cell(last_row, col)
            cell.alignment = Alignment(horizontal="left" if col in [3, 5, 10, 11] else "right")
            if last_row % 2 == 0:
                cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

        profit_margin = self.get_profit_margin() / 100.0
        cost = subtotal * (1 - profit_margin)
        profit = (subtotal - cost) - discount_total

        pl_data = [bill_no, date_str, subtotal, discount_total, grand_total, profit]
        ws_pl.append(pl_data)

        last_row_pl = ws_pl.max_row
        for col in range(1, 7):
            cell = ws_pl.cell(last_row_pl, col)
            cell.alignment = Alignment(horizontal="right" if col > 2 else "left")
            if last_row_pl % 2 == 0:
                cell.fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")

        wb.save(self.filename)
        return True

    # ---------- Retrieve Bills ----------
    def get_recent_bills(self, limit=10):
        wb = load_workbook(self.filename, data_only=True)
        ws = wb["Bills"]
        bills = []
        start_row = max(2, ws.max_row - limit + 1)
        for row in range(start_row, ws.max_row + 1):
            sr = ws.cell(row, 1).value
            if sr:
                bills.append({
                    'sr_no': sr,
                    'date': ws.cell(row, 2).value,
                    'customer_name': ws.cell(row, 3).value,
                    'customer_id': ws.cell(row, 4).value,
                    'bill_no': ws.cell(row, 5).value,
                    'items_count': ws.cell(row, 6).value,
                    'subtotal': float(ws.cell(row, 7).value or 0),
                    'discount': float(ws.cell(row, 8).value or 0),
                    'total': float(ws.cell(row, 9).value or 0),
                    'qr_file': ws.cell(row, 10).value,
                    'pdf_file': ws.cell(row, 11).value
                })
        return bills[::-1]

    def get_all_bills(self):
        return self.get_recent_bills(limit=100000)

    # ---------- Summary ----------
    def generate_summary_report(self):
        wb = load_workbook(self.filename, data_only=True)
        ws_pl = wb["ProfitLoss"]
        total_sales = 0
        total_discounts = 0
        total_profit = 0
        for row in range(2, ws_pl.max_row + 1):
            total_sales += ws_pl.cell(row, 3).value or 0
            total_discounts += ws_pl.cell(row, 4).value or 0
            total_profit += ws_pl.cell(row, 6).value or 0

        ws_bills = wb["Bills"]
        total_bills = ws_bills.max_row - 1
        return {
            'total_bills': max(0, total_bills),
            'total_sales': total_sales,
            'total_discounts': total_discounts,
            'total_profit': total_profit,
            'profit_margin': self.get_profit_margin()
        }

    def get_excel_file(self):
        return self.filename

    # ---------- Reset Methods ----------
    def erase_all_bills(self):
        try:
            wb = load_workbook(self.filename)
            ws_bills = wb["Bills"]
            if ws_bills.max_row > 1:
                ws_bills.delete_rows(2, ws_bills.max_row - 1)
            ws_pl = wb["ProfitLoss"]
            if ws_pl.max_row > 1:
                ws_pl.delete_rows(2, ws_pl.max_row - 1)
            wb.save(self.filename)
            return True
        except Exception as e:
            print(f"Error erasing bills: {e}")
            return False

    def factory_reset(self):
        try:
            wb = load_workbook(self.filename)
            ws_bills = wb["Bills"]
            if ws_bills.max_row > 1:
                ws_bills.delete_rows(2, ws_bills.max_row - 1)
            ws_pl = wb["ProfitLoss"]
            if ws_pl.max_row > 1:
                ws_pl.delete_rows(2, ws_pl.max_row - 1)
            ws_settings = wb["Settings"]
            for row in range(2, ws_settings.max_row + 1):
                key = ws_settings.cell(row, 1).value
                if key == "profit_margin":
                    ws_settings.cell(row, 2).value = 30.0
                elif key == "customer_counter":
                    ws_settings.cell(row, 2).value = 1000
                elif key == "bill_counter":
                    ws_settings.cell(row, 2).value = 1
                elif key == "store_name":
                    ws_settings.cell(row, 2).value = "My Store"
                elif key == "store_tagline":
                    ws_settings.cell(row, 2).value = "Quality You Can Trust"
                elif key == "store_address":
                    ws_settings.cell(row, 2).value = "123 Main Street, City - 000000"
                elif key == "store_phone":
                    ws_settings.cell(row, 2).value = "+91 98765 43210"
                elif key == "store_email":
                    ws_settings.cell(row, 2).value = "contact@mystore.com"
                elif key == "store_website":
                    ws_settings.cell(row, 2).value = "www.mystore.com"
                elif key == "currency_symbol":
                    ws_settings.cell(row, 2).value = "₹"
            wb.save(self.filename)
            return True
        except Exception as e:
            print(f"Error during factory reset: {e}")
            return False
